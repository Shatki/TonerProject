"""
    Библиотека для преобразования XLSX templates файлов в PDF для печати

"""
import os
from TonerProject.settings import BASE_DIR, STATIC_URL, DOCUMENT_DIR
from datetime import datetime

# openpyxl
from openpyxl import load_workbook
from openpyxl.utils.units import DEFAULT_ROW_HEIGHT, DEFAULT_COLUMN_WIDTH

# reportlab
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import A4, portrait, landscape
from reportlab.platypus import Table, TableStyle

# для шрифтов
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class Form(object):
    """ Класс для работы с шаблонами в формате Excel .
    Читает и подготавливает данные для экспорта в PDF.
    """
    # print(DEFAULT_ROW_HEIGHT, DEFAULT_COLUMN_WIDTH)
    PORTRAIT, LANDSCAPE = 'portrait', 'landscape'
    DEFAULT_CELL_HEIGHT = 11.0  # 15.0
    DEFAULT_CELL_WIDTH = 14.0  # 51.85
    DEFAULT_SHEET_ORIENTATION = PORTRAIT

    horizontal_align = {
        'right': 'RIGHT',
        'left': 'LEFT',
        'center': 'CENTER',
        None: 'None'
    }

    vertical_align = {
        'center': 'MIDDLE',
        'bottom': 'BOTTOM',
        'top': 'TOP',
        None: 'None'
    }

    cell_border = {
        'thin': 0.5,
        'medium': 2,
        'thick': 4,
    }

    ORIENTATION = {
        'landscape': landscape(A4),
        'portrait': portrait(A4),
    }

    def __init__(self, context, file_prefix=u'document'):
        """ Initialise an object of the class Form

        :param context: context data for create the document form
        :type context: list

        :param file_prefix: file name prefix of the PDF file
        :type file_prefix: str

        :rtype: None

        """
        self.context = context
        self.orientation = self.DEFAULT_SHEET_ORIENTATION
        self.workbook = ''
        self.worksheet = ''
        self.work_file = u'Empty'
        self.title = u'Form'
        self.width = 0
        self.height = 0

        self.pdf_width = 0
        self.pdf_height = 0
        self.pages = 0

        """ Настраиваем шрифт """
        # Draw things on the PDF. Here's where the PDF generation happens.
        # See the ReportLab documentation for the full list of functionality.
        pdfmetrics.registerFont(TTFont('Arial', '../fonts/Arial.ttf'))
        # pdfmetrics.registerFont(TTFont('Arial Bold', '../fonts/Arial_Cyr.ttf'))

        """ Настраиваем стили и данные для report lab """
        self.data = []
        self.heights = []
        self.widths = []
        self.style = TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e0e0e0')),  # для разработки
        ])

        """ создаём наименование PDF файла """
        self.pdf_file_name = file_prefix + datetime.now().strftime('-%Y-%m-%d-%H-%M-%S') + '.pdf'
        # создаем PDF файл документа для отдачи
        self.pdf_file = os.path.join(BASE_DIR, 'static', 'documents', self.pdf_file_name)
        self.logo = None

    def __str__(self):
        return self.pdf_file_name

    def __repr__(self):
        return u"<Form %s.%s>" % (self.title, self.work_file)

    def render(self):
        if self.data and self.heights and self.widths and self.orientation:
            self.document = canvas.Canvas(self.pdf_file, pagesize=self.orientation, bottomup=1)
            self.pdf_width, self.pdf_height = self.orientation
            self.table = Table(self.data, colWidths=self.widths, rowHeights=self.heights, repeatRows=1)
            self.table.setStyle(self.style)

            self.table_width, self.table_height = self.table.wrapOn(self.document, self.width, self.height)
            # print(table_height)
            self.table.drawOn(self.document, 0 + self.widths[0], 0 + self.heights[0], mm)
            self.document.showPage()
            self.document.save()

            # Информация для разработки
            print(self.worksheet.max_column, self.worksheet.max_row)
            print('Размеры PDF страницы(px): ', self.pdf_width, self.pdf_height)
            print('Размеры PDF страницы(mm): ', self.pdf_width / mm, self.pdf_height / mm)
            print('Размеры XLSX страницы(px): ', self.width, self.height)
            print('Размеры XLSX страницы(mm): ', self.width / mm, self.height / mm)

            # print(self.data)
            # print('width:', self.widths)
            # print('height:', self.heights)

            # my_cell = self.worksheet['AY6']
            # print('Ячейка:', my_cell.border.top.border_style)
            # print('Ячейки:', self.worksheet.merged_cells)
            # print('Объединены ячейки:', self.worksheet.merged_cell_ranges)

        else:
            print('Error: Load form first!')
            return None
        return True

    @classmethod
    def hexpr(cls, text):
        ret = ''
        for t in text:
            ret += hex(ord(t)) + '/'
        return ret

    @classmethod
    def wrap(cls, cell, context):
        """ Функция подготовки прочитанных значений из XLSX и обработки значений из контекста"""
        value = str(cell.value)
        br = chr(0x5c) + chr(0x6e)
        if value[0] == '#':
            #  Указание на повторяющуюся строку
            print('найден маркер строки: %s' % value[1:])
            return ''
        elif value[-2:] == '}}' and value[:2] == '{{':
            # Вставляем значение из контекста документа
            try:
                # На случай отсутствующих ключей
                value = context[value[2:-2]]
            except:
                print('Key Does not exist:', value[2:-2])
                value = ''
            return value
        # print(cell.value, cls.horisontal_align[cell.alignment.horizontal],
        #      cls.vertical_align[cell.alignment.vertical])
        # print(value, hex(ord('\\')), cls.hexpr(value))
        return value.replace(br, '\n')

    def load_xlsx(self, file_name, sheet_name=u'Sheet1'):
        """ Загрузка шаблона

        :type context: dict
        :param filename: Наименование файла
        :type filename: str

        :param sheetname: Наименование шаблона
        :type sheetname: str

        :rtype: tuple: (data, style, widths, heights)
        """
        self.work_file = file_name
        self.title = sheet_name
        self.workbook = load_workbook(filename=file_name)
        self.worksheet = self.workbook[sheet_name]
        # self.worksheet.cell.style.alignment.wrap_text = True
        """ считываем ориентацию страницы файла """
        self.orientation = self.ORIENTATION[self.worksheet.page_setup.orientation]


        """ Получаем размеры ячееек """
        first_row = self.worksheet[1]
        for cell in first_row:
            cel = self.worksheet.column_dimensions[cell.column]
            # print(cell.col_idx)
            if cel.hidden:
                # Скрытые столбцы
                self.widths.append(0)
            elif cel.width:
                self.widths.append(cel.width * 5.2)
                self.width += cel.width * 5.2
            else:
                self.widths.append(self.DEFAULT_CELL_WIDTH)
                self.width += self.DEFAULT_CELL_WIDTH

        for row in self.worksheet.rows:
            cell = self.worksheet.row_dimensions[row[0].row]
            # print(cell.hidden) Скрытые ячейки
            if cell.hidden:
                # Скрытые строки
                self.heights.append(0)
            elif cell.height:
                self.heights.append(cell.height)
                self.height += cell.height
            else:
                self.heights.append(self.DEFAULT_CELL_HEIGHT)
                self.height += self.DEFAULT_CELL_HEIGHT

        """ Получаем значения и стили из ячеек """
        #   Сформируем set() из всех ячеек
        for row in self.worksheet.rows:
            data_cells = []
            for cell in row:
                # if 7 <= cell.row <= 10:
                #    print('Ячейка: ', cell.coordinate, 'Border-',cell.border.top.border_style,
                #          cell.border.right.border_style, cell.border.bottom.border_style, cell.border.left.border_style,
                #          ' Еще:', cell.border.start, cell.border.end,
                #          cell.border.vertical, cell.border.horizontal)
                if cell.value:
                    data_cells.append(self.wrap(cell, self.context))
                    # размер шрифта
                    # print(cell.alignment.vertical)
                    # Размер шрифта
                    self.style.add('FONTSIZE', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                   cell.font.sz)
                    # Горизонтальное выравнивание текста
                    if cell.alignment.horizontal:
                        self.style.add('ALIGN', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                       self.horizontal_align[cell.alignment.horizontal])

                    # Вертикальное выравнивание текста
                    if cell.alignment.vertical:
                        self.style.add('VALIGN', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                       self.vertical_align[cell.alignment.vertical])

                        # print('Толщина:', cell.border.top.border_style)
                else:
                    data_cells.append('')

                    # if cell.coordinate not in self.worksheet.merged_cells:    # работает, но очень медленно
                    # Верхная граница
                if cell.border.top.border_style:
                    self.style.add('LINEABOVE', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                   self.cell_border[cell.border.top.border_style], colors.black)
                # Нижняя граница
                if cell.border.bottom.border_style:
                    self.style.add('LINEBELOW', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                   self.cell_border[cell.border.bottom.border_style], colors.black)
                # Левая граница
                if cell.border.left.border_style:
                    self.style.add('LINEBEFORE', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                   self.cell_border[cell.border.left.border_style], colors.black)
                # Правая граница
                if cell.border.right.border_style:
                    self.style.add('LINEAFTER', (cell.col_idx - 1, cell.row - 1), (cell.col_idx - 1, cell.row - 1),
                                   self.cell_border[cell.border.right.border_style], colors.black)

            self.data.append(data_cells)

        """ Формируем стили для объединённых ячеек """
        """ Make a style for merged cells """

        # for cell in self.worksheet['AY7:BF8']:
        #            print('Ячейка: ', cell, 'Border-')

        # print(self.worksheet._styles)

        ranges = []
        # Список объединений переведенных в цифровое обозначение
        for merged_range in self.worksheet.merged_cell_ranges:
            cell_range = merged_range.split(':', 2)
            cell_start = self.worksheet[cell_range[0]]  # used the new method: worksheet['A1']
            cell_end = self.worksheet[cell_range[1]]
            ranges.append([cell_start.col_idx - 1, cell_start.row - 1, cell_end.col_idx - 1, cell_end.row - 1])

            # Верхная граница
            if cell_start.border.top.border_style:
                # print('LINEABOVE', span[0], span[1], span[2], span[3])
                self.style.add('LINEABOVE',
                               (cell_start.col_idx - 1, cell_start.row - 1), (cell_end.col_idx - 1, cell_start.row - 1),
                               self.cell_border[cell_start.border.top.border_style], colors.black)
            # Левая граница
            if cell_start.border.left.border_style:
                self.style.add('LINEBEFORE',
                               (cell_start.col_idx - 1, cell_start.row - 1), (cell_start.col_idx - 1, cell_end.row - 1),
                               self.cell_border[cell_start.border.left.border_style], colors.black)

            # print(cell_start.border.bottom.border_style)
            # Нижняя граница
            if cell_start.border.bottom.border_style:
                self.style.add('LINEBELOW',
                               (cell_start.col_idx - 1, cell_end.row - 1), (cell_end.col_idx - 1, cell_end.row - 1),
                               self.cell_border[cell_start.border.bottom.border_style], colors.black)

            # Правая граница
            if cell_end.border.right.border_style:
                self.style.add('LINEAFTER',
                               (cell_end.col_idx - 1, cell_start.row - 1), (cell_end.col_idx - 1, cell_end.row - 1),
                               self.cell_border[cell_end.border.right.border_style], colors.black)

        for span in ranges:
            self.style.add(
                'SPAN', (span[0], span[1]), (span[2], span[3])
            )
        return True
